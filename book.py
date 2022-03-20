# -*- coding: utf8 -*- 
# step1. 관련 패키지 및 모듈 불러오기

# 이미지는 가져오지 않음

from re import I
from selenium import webdriver
import time
from bs4  import BeautifulSoup as bs
import openpyxl
import platform
from openpyxl.styles import Alignment, Border, Side
import re


thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))


# step2. 네이버 뉴스 댓글정보 수집 함수
def get_book_site(url, wait_time=5, delay_time=0.1):

    wb = openpyxl.Workbook()
    ws = wb.active
    # 크롬 드라이버로 해당 url에 접속
    if platform.system() == "Windows":
        driver = webdriver.Chrome("D:/7.Software/chromedriver_win32/chromedriver")
    else:
        driver = webdriver.Chrome("/home/lima/Work_Python/chromedriver_linux64/chromedriver")
    
        # (크롬)드라이버가 요소를 찾는데에 최대 wait_time 초까지 기다림 (함수 사용 시 설정 가능하며 기본값은 5초)
    driver.implicitly_wait(wait_time)
    
    # 인자로 입력받은 url 주소를 가져와서 접속
    driver.get(url)

    endbtn = driver.find_element_by_class_name('bgYUI.end')
    endnum = int(endbtn.get_attribute('title'))
    endnum = 20 if endnum > 20 else endnum
    for i, title in enumerate(['번호', '책명', '출판사', '출시일', '판매가격', '판매지수', '평가']):
        ws.cell(1, i+1).value = title
        ws.cell(row=1, column=i+1).alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        ws.cell(row=1, column=i+1).border = thin_border

    row = 2    
    for i in range(1, endnum+2):
        driver.get(url + str(i))
        searchlist=driver.find_element_by_id('yesSchList')
        item_info = searchlist.find_elements_by_css_selector('div.item_info')

        for gd in item_info:
            # page=gd.get_attribute('innerHTML')
            # print(page)
            ws.cell(row, 1).value = row - 1
            try:
                info_name=gd.find_element_by_class_name('gd_name')
                ws.cell(row, 2).value = info_name.text
                print(row, info_name.text)
            except:
                print(i, 'error:info_name')
                pass

            

            try:
                info_pub = gd.find_element_by_class_name('authPub.info_pub')
                ws.cell(row, 3).value = info_pub.text
            except:
                print(i, 'error:info_pub')
                pass

            
            
            try:
                info_date = gd.find_element_by_class_name('authPub.info_date')
                ws.cell(row, 4).value = info_date.text
            except:
                print(i, 'error:info_date')
                pass

            try:
                info_price=gd.find_element_by_css_selector('div.info_row.info_price > strong > em')
                ws.cell(row, 5).value = info_price.text
                print(row, info_price.text)
            except:
                print(i, 'error:info_name')
                pass

            try:
                saleNum=gd.find_element_by_class_name('saleNum')
                ws.cell(row, 6).value = re.search(r'([0-9,]+)', saleNum.text)[1]
                print(re.search(r'([0-9,]+)', saleNum.text)[1])
            except:
                print(i, 'error:info_rating')
                pass

            try:
                ##yesSchList > li:nth-child(1) > div > div.item_info > div.info_row.info_rating > span.rating_grade > em
                yes_b=gd.find_element_by_css_selector('span.rating_grade > em')
                ws.cell(row, 7).value = re.search(r'([0-9.]+)', yes_b.text)[1]
                print(re.search(r'([0-9.]+)', yes_b.text)[1])
            except:
                print(i, 'error:info_rating')
                pass


            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='center')
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            ws.cell(row=row, column=7).border = thin_border

            row += 1

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        wb.save('./book.xlsx')
    wb.close

if __name__ == '__main__':
    # in_str=input("input the search word:")
    searchword = '초등수학'

    url = 'http://www.yes24.com/Product/Search?domain=ALL&query=' + searchword + '&page='
    get_book_site(url)