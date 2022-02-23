
"Global-Autonews Crawling"
from types import NoneType
from urllib.parse import quote_plus
from bs4 import BeautifulSoup
from selenium import webdriver
import openpyxl
from openpyxl.styles import PatternFill
import urllib.request as req
import datetime
from datetime import date
import time
from time import mktime
from urllib.error import URLError, HTTPError
import re
from urllib.request import Request, urlopen


def read_url(url):
    try:       
        headers = {'User-Agent': 'Mozilla/5.0'}
        request = Request(url, headers=headers)
        res = urlopen(request)
        time.sleep(0.3)
        html = res.read()
        soup = BeautifulSoup(html,'html.parser')
    except HTTPError as e:
        err = e.read()
        code = e.getcode()
        print(err)
        print(code)
    return soup


def read_url_with_login(url, id, pd):
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    # 아래 로그들이 출력되는 것을 방지 하기 위함
    #DevTools listening on ws://127.0.0.1:54448/devtools/browser/e1ea556b-0e77-453c-9875-08d81ae3383b

    driver = webdriver.Chrome('../chromedriver_win32/chromedriver.exe', options=options)
    driver.get(url)
    
    driver.find_element_by_name('profiles.login.login_id').send_keys(id)
    driver.find_element_by_name('profiles.login.password').send_keys(pd)

    driver.find_element_by_id('login_btn').click()
    
    driver.implicitly_wait(10)
    
    cur_url = driver.current_url
    driver.get(cur_url) #changed here
    res = driver.page_source
    soup = BeautifulSoup(res,'html.parser')
    driver.close()
    return soup

  
def read_subtitle(suburl):
    soup = read_url(suburl)
    subtitle = soup.select_one('h2.article-sub-title').text    
    return subtitle

def read_subtitle2(suburl):
    soup = read_url(suburl)
    subtitle = soup.select_one('h2.preview').text    
    return subtitle

def read_subtitle3(suburl):
    # print(suburl)
    soup = read_url(suburl)
    # f = open('D:/url.html', 'w', encoding='UTF-8')
    # f.write(suburl)
    # f.write(str(soup))
    # f.close()
    
    # stamp = soup.select_one('#main_section > div > time')
    #stamp = soup.select_one('div.details > time')
    stamp_str = soup.select_one('time.updated')
    if stamp_str:
        # print(stamp_str)
        stamp = re.search(r"([0-9a-zA-Z, ]+)", stamp_str.text)
        return 1, stamp.group(0)
    else:
        #report_detail > div.col-md-12 > div:nth-child(2) > h3.page-updateAt
        stamp_str = soup.select_one('div > div > h3.page-updateAt')
        # print(stamp_str)
        stamp = re.search(r"([0-9a-zA-Z/ ]+)", stamp_str.text)
        return 2, stamp.group(0)

    
def read_globalauto():
    url = "http://global-autonews.com/bbs/board.php?bo_table=bd_001"
    soup = read_url(url)
    
    titles = soup.select('div[class=list_box]')
    for title in titles:
        content_date = title.select_one('span[class=date]').text.replace('\n', '')
        
        if datetime.datetime.strptime(content_date, "%Y-%m-%d").date() >= now - datetime.timedelta(days=1) :
            subject=title.select_one('label[class=sound_only]').text.replace('\t','')
            content =title.select_one('span[class=list_content]').text.replace('\t','')
            link=title.select('a')[0]['href']
            news_list.append([len(news_list), "Global Auto News", content_date, subject, content, link])


def read_auto():
    url2= "https://www.autonews.com/manufacturing"
    soup = read_url(url2)
    
    titles = soup.select('div.views-row')
    for title in titles:
        time_str = title.select_one('span.article-update-time')
        gmt_time=re.search(r".*--(\d+).*", time_str.attrs['data-lastupdated']).group(1)
        local_time = time.localtime(int(gmt_time))
        content_date = datetime.datetime.fromtimestamp(mktime(local_time)).date()
        
        if content_date >= now - datetime.timedelta(days=1) :
            subject=title.select_one('div.feature-article-headline').text
            link="https://www.autonews.com" + title.select('a')[0]['href']
            content = read_subtitle(link)
            news_list.append([len(news_list),"Automotive News", content_date, subject, content, link])


def read_insideevs():
    url2= "https://insideevs.com/news/"
    soup = read_url(url2)
    
    titles = soup.select('div.item.wcom')
    for title in titles:
        time_str=title.select_one('span.date')
        gmt_time=re.search(r"(\d+)", time_str.attrs['data-time']).group(1)
        
        local_time = time.localtime(int(gmt_time))
        content_date = datetime.datetime.fromtimestamp(mktime(local_time)).date()
        
        if content_date >= now - datetime.timedelta(days=1) :
            subject=title.select_one('h3').text
            link="https://insideevs.com" + title.select('a')[0]['href']
            content = read_subtitle2(link)
            news_list.append([len(news_list),"InsideEVs", content_date, subject, content, link])


       
def read_marklines():
    url2= "https://www.marklines.com/en/members/login"
    soup = read_url_with_login(url2, "", "")
    
    ##new_topics_1 > div > div.col-md-3 > div > div > a:nth-child(1)
    new_topics = soup.select('div.ml_new_topics > div > a')
    # print(new_topics)
    for new_topic in new_topics:
        link = "https://www.marklines.com" + new_topic['href']
        type, time_str = read_subtitle3(link)
        if type == 1:
            content_date = datetime.datetime.strptime(time_str, "%b %d, %Y").date()
        elif type == 2:
            content_date = datetime.datetime.strptime(time_str, "%Y/%m/%d").date()
        
        if content_date >= now - datetime.timedelta(days=1) :
            subject = new_topic.text.strip()
            content = ""
            news_list.append([len(news_list),"Marklines", str(content_date), subject, content, link])
            # print("Marklines", str(content_date), subject, content, link)
    #new_topics_1 > div > div.col-md-9.topics > div:nth-child(7) > a
    topics = soup.select('div.col-md-9.topics > div > a')
    for topic in topics:
        # print(topic)
        link = "https://www.marklines.com" + topic['href']
        type, time_str = read_subtitle3(link)
        if type == 1:
            content_date = datetime.datetime.strptime(time_str, "%b %d, %Y").date()
        elif type == 2:
            content_date = datetime.datetime.strptime(time_str, "%Y/%m/%d").date()

        if content_date >= now - datetime.timedelta(days=1) :
            subject = topic.text.strip()
            content = ""
            news_list.append([len(news_list),"Marklines", str(content_date), subject, content, link])
            # print("Marklines", str(content_date), subject, content, link)
            
def write_excel(data):
    from openpyxl.styles import Alignment
    from openpyxl.styles import Border, Side
    
    filepath = "D:\\news4.xlsx"
    wb = openpyxl.Workbook()
    # 현재 활성 중인 워크시트를 선택하는 방법은 다음과 같습니다.
    sheet = wb.active
    sheet.title = "news"
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    thick_border = Border(left=Side(style='thick'), 
                         right=Side(style='thick'), 
                         top=Side(style='thick'), 
                         bottom=Side(style='thick'))
    
    for i, d1 in enumerate(data):
        for j, d2 in enumerate(d1):
            # print(i,j, d2)
            c2 = sheet.cell(row=i+1, column=j+1)
            if j==0 and i>0:
                c2.value = i
            else:
                c2.value = d2
                
            if i==0:
                c2.border = thick_border
            else:
                c2.border = thin_border

    for row in sheet[1:sheet.max_row]:
        cell = row[1]             # column B
        cell.alignment = Alignment(horizontal='center')
        cell = row[2]             # column B
        cell.alignment = Alignment(horizontal='center')
    
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 80
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 30

    wb.save(filepath)

import pickle

def debug_read():
    with open('D:/parrot.pkl', 'rb') as p:
        mynewlist = pickle.load(p)
        write_excel(mynewlist)

if __name__ == "__main__":
    
    with open('D:/parrot.pkl', 'wb') as p:
        now = date.today() 
        news_list = []#["No", "Category", "Date", "Headline", "Summary", "Link"]

        read_globalauto()
        read_auto()
        read_insideevs()
        read_marklines()
        
        # s = sorted(news_list, key = lambda x: (x[0], -x[1]))
        s = sorted(news_list, key=lambda x: (-ord((x[1].lower())[0]), x[2]), reverse=True)

        write_excel(s)
        s.insert(0, ["No", "Category", "Date", "Headline", "Summary", "Link"])
        for news in s:
            print(news)
            
        pickle.dump(s, p)
