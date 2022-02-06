
from importlib.resources import open_binary
from logging import INFO
import os
import sys
import time
import datetime
from turtle import onclick

import requests
from bs4 import BeautifulSoup as bs
import dateutil.relativedelta
import re

import json
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def make_golf_address():
    sess = requests.session()

    info = {
        'p_Type': 'GolfCourseTableGrid',
        'p_SEARCH_HOLE_CN': '',
        'p_SEARCH_LOCATION': '',
        'p_SEARCH_TEXT': '',
        'p_PAGE_NO': 1,
        'p_LIST_CN': 100
    }
    url = 'http://www.kgagolf.or.kr/GolfCourse/DataService/JqCommonContents.aspx'

    page = sess.post(url, info)
    soup = bs(page.content, 'html.parser')
    print(soup.text)

    jsonObject = json.loads(soup.text)
    jsonArray = jsonObject.get("rows")

    wb = Workbook()
    sheet = wb.active

    title = ['No','Name','Hole','Address','Tel','Fax']
    for i, ti in enumerate(title):
        column_letter = get_column_letter(i+1)
        ind = '%s%d' % (column_letter, 1)
        sheet[ind] = ti

    for row, ii in enumerate(jsonArray):
        for col, item in enumerate(ii['row']):
            column_letter = get_column_letter(col+1)
            print(item.get("Text"))
            ind = '%s%d' % (column_letter, row + 2)
            sheet[ind] = item.get("Text")
            # print(item)
            # golf=dict(str(item))
            # print(golf["Text"])

    wb.save('golf.xlsx')


if __name__ == '__main__':
    make_golf_address()