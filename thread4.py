import sys
import PyQt5
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5 import uic

# This is our window from QtCreator
# import mainwindow_auto
form_class = uic.loadUiType("/home/lima/Work_Python/Study/study.ui")[0]
from openpyxl import load_workbook
import re
from PyQt5.QtCore import Qt

class myExcel():
    def __init__(self, filepath=None):
        if filepath != None:
            self.row = 1
            self.wb  = load_workbook('/home/lima/Work_Python/Scraping/exam.xlsx')
            self.ws  = self.wb.active
        else:
            self.row = 0
            self.wb  = None#load_workbook('/home/lima/Work_Python/Scraping/exam.xlsx')
            self.ws  = None#self.wb.active

    def open_excel(self, filepath):
        if self.wb != None:
            self.wb.close()

        self.row = 1
        self.wb  = load_workbook(filepath)
        self.ws  = self.wb.active

    def get_total(self):
        print(self.ws.max_row)
        return self.ws.max_row

    def read_tab(self):
        for tab in self.wb.sheetnames:
            print(tab)

        res = sorted(map(str, self.wb.sheetnames), key = lambda ele: (0, int(re.search(r'(\d+).*', ele)[1]))
                        if re.search(r'(\d+).*', ele) else (1, ele))
        return res

    def select_sheet(self, name):
        print('sheet select', name)
        self.ws = self.wb[name]
        self.row = 1

    def read_current(self, inc=0):
        print('excel_read', self.row)
        if self.row >= 1 and self.row <= self.ws.max_row:
            data = [cell.value for cell in self.ws[self.row]]

        if self.row + inc <= self.ws.max_row and self.row + inc >= 1:
            self.row += inc

        return data

    def set_row(self, index):
        if self.row >= 1 and self.row <= self.ws.max_row:
            self.row = index
        data = [cell.value for cell in self.ws[self.row]]
        return data

    def get_row(self):
        return self.row

    def read_next(self):
        print('next', self.row)
        
        if self.row < self.ws.max_row:
            self.row += 1

        data = [cell.value for cell in self.ws[self.row]]
        return data

    def read_prev(self):
        print('prev', self.row)
        
        if self.row > 1:
            self.row -= 1

        data = [cell.value for cell in self.ws[self.row]]
        return data


class DataCaptureThread(QThread):
    def collectProcessData(self):
        print ("Collecting Process Data")
        data = self.excel.read_current(1)
        self.parent.updateQnA(data[0], data[1])
        self.parent.progressBar.setValue(self.excel.get_row())

    def __init__(self, parent, excel):
        QThread.__init__(self, parent)
        self.parent = parent
        self.dataCollectionTimer = PyQt5.QtCore.QTimer()
        self.dataCollectionTimer.moveToThread(self)
        self.dataCollectionTimer.timeout.connect(self.collectProcessData)
        self.excel = excel
        self.collectProcessData()

    def run(self):
        print('Thread run')
        duration = int(self.parent.lineEdit_time.text())
        if duration >=1 and duration <100:
            self.duration = duration
        else:
            self.duration = 5
        self.dataCollectionTimer.start(self.duration*1000)
        loop = PyQt5.QtCore.QEventLoop()
        loop.exec_()

default_file = '/home/lima/Work_Python/Scraping/exam.xlsx'
class MainWindow(QMainWindow, form_class):
    def __init__(self):
        super(self.__class__, self).__init__()
        self.setupUi(self) # gets defined in the UI file
        self.runButton1.clicked.connect(self.pressedStartBtn)
        self.runButton2.clicked.connect(self.pressedStopBtn)
        self.Button_prev.clicked.connect(self.pressedPrevBtn)
        self.Button_next.clicked.connect(self.pressedNextBtn)
        self.Button_answer.clicked.connect(self.pressedAnswerBtn)

        self.excel = myExcel(default_file)
        for tab in self.excel.read_tab():
            self.comboBox_episode.addItem(tab)
        
        
        # self.comboBox_episode.currentIndexChanged().connect(self.onTabChanged)
        self.onTabChanged(self.comboBox_episode.currentText())
        self.comboBox_episode.activated[str].connect(self.onTabChanged)
        self.dataCollectionThread = None
        self.auto = False
        self.pushButton_openfile.clicked.connect(self.fileOpen)
        self.lineEdit_inputfile.setText(default_file)
        self.actionOpen.triggered.connect(self.fileOpen)
        self.actionStay_On.triggered.connect(self.windowDisplay)


    def updateQnA(self, question, answer):
        self.textBrowser_question.clear()
        self.textBrowser_answer.clear()
        self.textBrowser_question.append(question)
        self.textBrowser_question.moveCursor(1)

        if self.checkBox_answer.isChecked():
            self.textBrowser_answer.append(answer)
            self.textBrowser_answer.moveCursor(1)

    def onTabChanged(self, value):
        self.excel.select_sheet(value)
        data = self.excel.read_current(0)
        self.updateQnA(data[0], data[1])
        self.progressBar.setMaximum(self.excel.get_total())
        self.progressBar.setValue(self.excel.get_row())
        

    def pressedStartBtn(self):
        # self.lblAction.setText("STARTED")
        self.auto = True
        self.dataCollectionThread = DataCaptureThread(self, self.excel)
        self.dataCollectionThread.start()

    def pressedStopBtn(self):
        # self.lblAction.setText("STOPPED")
        self.auto = False
        if self.dataCollectionThread != None:
            self.dataCollectionThread.terminate()
            self.dataCollectionThread = None
            self.excel.set_row(self.excel.get_row()-1)

    def pressedPrevBtn(self):
        self.pressedStopBtn()
        data=self.excel.read_prev()
        self.updateQnA(data[0], data[1])
        self.progressBar.setValue(self.excel.get_row())

    def pressedNextBtn(self):
        self.pressedStopBtn()
        data=self.excel.read_next()
        self.updateQnA(data[0], data[1])
        self.progressBar.setValue(self.excel.get_row())

    def pressedAnswerBtn(self):
        pass

    def fileOpen(self):
        self.fname = QFileDialog.getOpenFileName(self, 'Open file', './')[0]
        self.lineEdit_inputfile.setText(self.fname)
        self.excel.open_excel(self.fname)
        self.comboBox_episode.clear()
        for tab in self.excel.read_tab():
            self.comboBox_episode.addItem(tab)
        self.onTabChanged(self.comboBox_episode.currentText())

    def windowDisplay(self):
        on = bool(self.windowFlags() & Qt.WindowStaysOnTopHint)
        print(on)
        # toggle the state of the flag
        self.setWindowFlag(Qt.WindowStaysOnTopHint)
        self.show()

def main():
     # a new app instance
     app = QApplication(sys.argv)
     form = MainWindow()
     form.show()
     sys.exit(app.exec_())

if __name__ == "__main__":
     main()