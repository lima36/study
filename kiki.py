# -*- coding: utf8 -*- 
# step1. 관련 패키지 및 모듈 불러오기

# 이미지는 가져오지 않음

from selenium import webdriver
import time
from bs4  import BeautifulSoup as bs
import openpyxl
import platform
from openpyxl.styles import Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active

thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

# step2. 네이버 뉴스 댓글정보 수집 함수
def get_naver_news_comments(url, wait_time=5, delay_time=0.1):
    # 크롬 드라이버로 해당 url에 접속
    if platform.system() == "Windows":
        driver = webdriver.Chrome('D:/Software/chromedriver')
    else:
        driver = webdriver.Chrome("/home/lima/Work_Python/chromedriver_linux64/chromedriver")
    
    # (크롬)드라이버가 요소를 찾는데에 최대 wait_time 초까지 기다림 (함수 사용 시 설정 가능하며 기본값은 5초)
    driver.implicitly_wait(wait_time)
    
    # 인자로 입력받은 url 주소를 가져와서 접속
    driver.get(url)

    element = driver.find_element_by_id("cafe_main")
    driver.switch_to.frame(element)
    
    # 댓글로 이동
    comments=driver.find_element_by_css_selector('a.button_comment')
    comments.click()
    time.sleep(10)

    nextbtns = driver.find_elements_by_css_selector('div.CommentBox > div.ArticlePaginate > button')
    ind = 1
    for i, next in enumerate(nextbtns):
        # Page 이동
        nextbtns[i].click()
        
        # 본격적인 크롤링 타임
        commentlist = driver.find_element_by_css_selector("ul[class='comment_list'")
        html = commentlist.get_attribute('innerHTML')
        soup = bs(html, 'html.parser')
        j=5
        for li in soup.select('li'):
            if "CommentItem--reply" in li.attrs['class']:
                # print(li.text.strip())
                
                nickname=li.find('div', class_='comment_nick_info')
                textbox = li.find('div', class_='comment_text_box')
                textdate = li.find('span', class_='comment_info_date')
                
                # CommentItem 에서 index를 1로 증가 시키기 전의 댓글로 등록되어야 함.
                ws.cell(ind-1, j).value = str(textdate.text.strip())
                ws.cell(ind-1, j+1).value = str(nickname.text.strip())
                ws.cell(ind-1, j+2).value = str(textbox.text.strip())
                    
                ws.cell(ind-1, j).alignment = Alignment(wrap_text=True, vertical='top')
                ws.cell(ind-1, j).border = thin_border
                ws.cell(ind-1, j+1).alignment = Alignment(wrap_text=True, vertical='top')
                ws.cell(ind-1, j+1).border = thin_border
                ws.cell(ind-1, j+2).alignment = Alignment(wrap_text=True, vertical='top')
                ws.cell(ind-1, j+2).border = thin_border

                j += 3
            elif "CommentItem" in li.attrs['class']:
                j = 5
                deleted = li.find('p', class_='comment_deleted')
                if deleted:
                    print("comment deleted")
                    continue
                try:
                    id = li.attrs['id'].strip()
                    nickname=li.find('div', class_='comment_nick_info')
                    textbox = li.find('div', class_='comment_text_box')
                    textdate = li.find('span', class_='comment_info_date')
                except:
                    print(i, ind, li)
                    print('error')
                    continue

                # print(nickname.text.strip())
                # print(id, textbox.text.strip())
                # print(id, textdate.text.strip())
                
                ws.cell(ind, 1).value = str(ind)
                ws.cell(ind, 2).value = str(textdate.text.strip())

                ws.cell(ind, 3).value = str(nickname.text.strip())
                ws.cell(ind, 4).value = str(textbox.text.strip())
                
                ws.cell(row=ind, column=1).alignment = Alignment(wrap_text=True, vertical='top')
                ws.cell(row=ind, column=1).border = thin_border
                ws.cell(row=ind, column=2).alignment = Alignment(wrap_text=True, vertical='top') 
                ws.cell(row=ind, column=2).border = thin_border
                ws.cell(row=ind, column=3).alignment = Alignment(wrap_text=True, vertical='top')
                ws.cell(row=ind, column=3).border = thin_border
                ws.cell(row=ind, column=4).alignment = Alignment(wrap_text=True, vertical='top')
                ws.cell(row=ind, column=4).border = thin_border
                
                ind += 1
            
        time.sleep(10)
        print(next.text)

        ws.column_dimensions['A'].width = 5 # index
        ws.column_dimensions['B'].width = 10 # time
        ws.column_dimensions['C'].width = 10 # nickname
        ws.column_dimensions['D'].width = 80 # text
        ws.column_dimensions['E'].width = 10 # time
        ws.column_dimensions['F'].width = 10 # nickname
        ws.column_dimensions['G'].width = 80 # text
        ws.column_dimensions['H'].width = 10 # time
        ws.column_dimensions['I'].width = 10 # nickname
        ws.column_dimensions['J'].width = 80 # text
    
        wb.save('./kiki5.xlsx')
        print('saving')

    wb.close()


# step3. 실제 함수 실행 및 엑셀로 저장
if __name__ == '__main__': # 설명하자면 매우 길어져서 그냥 이렇게 사용하는 것을 권장

    # 원하는 기사 url 입력
    # url = '댓글 크롤링 원하는 기사의 url (주의! 댓글보기를 클릭한 상태의 url이어야 함'
    url = 'https://cafe.naver.com/dochithink?iframe_url_utf8=%2FArticleRead.nhn%253Fclubid%3D12843510%2526articleid%3D1832183%2526referrerAllArticles%3Dtrue'
    # 함수 실행
    get_naver_news_comments(url)
